"""
AutoTestDesign - ExportAgent (FR6.0)
将测试产物序列化为 JSON / Excel / CSV 格式
"""

import os
import json
import csv
from datetime import datetime
from typing import List, Dict, Any

from graph.state import AutoTestState

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ─────────────────────────────────────────────
# 辅助函数
# ─────────────────────────────────────────────

def safe_str(value) -> str:
    if isinstance(value, list):
        return "\n".join(str(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value) if value is not None else ""


def get_output_dir() -> str:
    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "outputs")
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


def timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ─────────────────────────────────────────────
# JSON 导出
# ─────────────────────────────────────────────

def export_json(state: AutoTestState, output_dir: str) -> str:
    payload = {
        "export_timestamp": datetime.now().isoformat(),
        "summary": {
            "requirements_count": len(state.get("structured_requirements", [])),
            "risk_items_count": len(state.get("risk_analysis", [])),
            "blackbox_tests_count": len(state.get("blackbox_tests", [])),
            "whitebox_tests_count": len(state.get("whitebox_tests", [])),
            "optimized_suite_count": len(state.get("optimized_suite", [])),
        },
        "requirements": state.get("structured_requirements", []),
        "risk_analysis": state.get("risk_analysis", []),
        "blackbox_tests": state.get("blackbox_tests", []),
        "whitebox_tests": state.get("whitebox_tests", []),
        "state_models": state.get("state_models", []),
        "optimized_suite": state.get("optimized_suite", []),
        "optimization_result": state.get("optimization_result", {}),
    }

    path = os.path.join(output_dir, f"test_suite_{timestamp()}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return path


# ─────────────────────────────────────────────
# CSV 导出
# ─────────────────────────────────────────────

def export_csv(state: AutoTestState, output_dir: str) -> str:
    """导出优化后的测试套件为 CSV"""
    test_cases = state.get("optimized_suite", []) or state.get("enriched_tests", [])
    path = os.path.join(output_dir, f"test_cases_{timestamp()}.csv")

    fieldnames = [
        "tc_id", "req_id", "technique", "title", "priority",
        "is_positive", "preconditions", "test_data",
        "expected_result", "coverage_tags"
    ]

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for tc in test_cases:
            row = {k: safe_str(tc.get(k, "")) for k in fieldnames}
            writer.writerow(row)

    return path


# ─────────────────────────────────────────────
# Excel 导出
# ─────────────────────────────────────────────

def style_header_row(ws, row_num: int, fill_color: str = "4472C4"):
    """为标题行设置样式"""
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", name="微软雅黑", size=10)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[row_num]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment


def auto_column_width(ws, min_width=10, max_width=50):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                cell_len = len(str(cell.value or ""))
                # CJK 字符算 2 倍宽度
                cjk_count = sum(1 for c in str(cell.value or "") if '\u4e00' <= c <= '\u9fff')
                adjusted_len = cell_len + cjk_count
                max_length = max(max_length, adjusted_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_length + 2))


def create_requirements_sheet(wb, structured_reqs: List[Dict]):
    ws = wb.create_sheet("需求分析")
    headers = ["需求ID", "标题", "描述", "领域", "输入字段", "数据范围", "条件", "预期行为"]
    ws.append(headers)
    style_header_row(ws, 1, "4472C4")

    for req in structured_reqs:
        ws.append([
            req.get("req_id", ""),
            req.get("title", ""),
            req.get("description", ""),
            req.get("domain", ""),
            safe_str([f.get("name", "") + "(" + f.get("data_type", "") + ")"
                      for f in req.get("input_fields", [])]),
            safe_str([f"{r.get('field_name', '')}:[{r.get('min_value', '')}~{r.get('max_value', '')}]"
                      for r in req.get("data_ranges", [])]),
            safe_str(req.get("conditions", [])),
            safe_str(req.get("expected_actions", [])),
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    auto_column_width(ws)
    ws.freeze_panes = "A2"


def create_risk_sheet(wb, risk_analysis: List[Dict]):
    ws = wb.create_sheet("风险分析报告")
    headers = ["需求ID", "标题", "影响度", "缺陷概率", "复杂度", "风险分数", "优先级", "风险因素", "缓解建议"]
    ws.append(headers)
    style_header_row(ws, 1, "C00000")

    priority_colors = {"High": "FF0000", "Medium": "FF9900", "Low": "70AD47"}

    for risk in risk_analysis:
        priority = risk.get("priority", "Medium")
        ws.append([
            risk.get("req_id", ""),
            risk.get("title", ""),
            risk.get("impact", ""),
            risk.get("probability", ""),
            risk.get("complexity", ""),
            risk.get("risk_score", ""),
            priority,
            safe_str(risk.get("risk_factors", [])),
            risk.get("mitigation", ""),
        ])
        # 颜色标记优先级列
        row_num = ws.max_row
        priority_cell = ws.cell(row=row_num, column=7)
        color = priority_colors.get(priority, "FFFFFF")
        priority_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        priority_cell.font = Font(bold=True, color="FFFFFF")

    auto_column_width(ws)
    ws.freeze_panes = "A2"


def create_test_cases_sheet(wb, test_cases: List[Dict], sheet_name: str, header_color: str):
    ws = wb.create_sheet(sheet_name)
    headers = [
        "测试用例ID", "需求ID", "技术方法", "标题", "优先级",
        "测试类型", "前置条件", "测试数据", "测试步骤", "预期结果", "覆盖标签"
    ]
    ws.append(headers)
    style_header_row(ws, 1, header_color)

    technique_map = {
        "Equivalence_Partitioning": "等价类划分",
        "Boundary_Value_Analysis": "边界值分析",
        "Decision_Table": "决策表",
        "State_Transition": "状态转换",
    }

    for tc in test_cases:
        steps_str = safe_str([
            f"{s.get('step_number', '')}.{s.get('action', '')} → {s.get('expected', '')}"
            for s in tc.get("test_steps", [])
        ])
        ws.append([
            tc.get("tc_id", ""),
            tc.get("req_id", ""),
            technique_map.get(tc.get("technique", ""), tc.get("technique", "")),
            tc.get("title", ""),
            tc.get("priority", ""),
            "正向" if tc.get("is_positive", True) else "负向",
            safe_str(tc.get("preconditions", [])),
            safe_str(tc.get("test_data", {})),
            steps_str,
            tc.get("expected_result", ""),
            safe_str(tc.get("coverage_tags", [])),
        ])
        # 颜色标记
        row_num = ws.max_row
        priority_colors = {"High": "FF0000", "Medium": "FF9900", "Low": "70AD47"}
        priority = tc.get("priority", "Medium")
        ws.cell(row=row_num, column=5).fill = PatternFill(
            start_color=priority_colors.get(priority, "FFFFFF"),
            end_color=priority_colors.get(priority, "FFFFFF"),
            fill_type="solid"
        )
        ws.cell(row=row_num, column=5).font = Font(bold=True, color="FFFFFF")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    auto_column_width(ws)
    ws.freeze_panes = "A2"


def create_summary_sheet(wb, state: AutoTestState):
    ws = wb.create_sheet("摘要总览", 0)  # 插入为第一个sheet

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    title_font = Font(bold=True, size=16, color="FFFFFF", name="微软雅黑")
    title_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    ws["A1"] = "AutoTestDesign - 测试套件报告"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    opt_result = state.get("optimization_result", {})
    data = [
        ("", ""),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("📋 需求分析", ""),
        ("  需求总数", len(state.get("structured_requirements", []))),
        ("", ""),
        ("⚠️ 风险分析", ""),
        ("  高风险需求", sum(1 for r in state.get("risk_analysis", []) if r.get("priority") == "High")),
        ("  中风险需求", sum(1 for r in state.get("risk_analysis", []) if r.get("priority") == "Medium")),
        ("  低风险需求", sum(1 for r in state.get("risk_analysis", []) if r.get("priority") == "Low")),
        ("", ""),
        ("🧪 测试用例生成", ""),
        ("  黑盒测试用例（EP+BVA+决策表）", len(state.get("blackbox_tests", []))),
        ("  白盒测试用例（状态转换）", len(state.get("whitebox_tests", []))),
        ("  合计", len(state.get("blackbox_tests", [])) + len(state.get("whitebox_tests", []))),
        ("", ""),
        ("✅ 优化结果", ""),
        ("  优化前用例数", opt_result.get("original_count", "N/A")),
        ("  优化后用例数", opt_result.get("optimized_count", "N/A")),
        ("  需求覆盖率", f"{opt_result.get('coverage_rate', 0):.1%}" if opt_result.get("coverage_rate") else "N/A"),
    ]

    for label, value in data:
        ws.append([label, value])
        if label and not label.startswith(" ") and label not in ("生成时间",):
            row = ws.max_row
            ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="2F5496")

    ws.freeze_panes = None


def export_excel(state: AutoTestState, output_dir: str) -> str:
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl 未安装，无法导出 Excel")

    wb = openpyxl.Workbook()
    # 删除默认 sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 创建各 sheet
    create_summary_sheet(wb, state)
    create_requirements_sheet(wb, state.get("structured_requirements", []))
    create_risk_sheet(wb, state.get("risk_analysis", []))
    create_test_cases_sheet(wb, state.get("blackbox_tests", []), "黑盒测试用例", "4472C4")
    create_test_cases_sheet(wb, state.get("whitebox_tests", []), "白盒测试用例", "7030A0")
    create_test_cases_sheet(wb, state.get("optimized_suite", []), "优化后测试套件", "375623")

    path = os.path.join(output_dir, f"AutoTestDesign_{timestamp()}.xlsx")
    wb.save(path)
    return path


# ─────────────────────────────────────────────
# 主节点函数
# ─────────────────────────────────────────────

def exporter_node(state: AutoTestState) -> Dict[str, Any]:
    """
    LangGraph 节点函数
    全部测试产物 → 导出文件
    """
    output_dir = get_output_dir()
    progress = ["[Exporter] 开始导出测试产物..."]
    errors = []
    paths = {}

    # 导出 JSON
    try:
        json_path = export_json(state, output_dir)
        paths["json_path"] = json_path
        progress.append(f"[Exporter] JSON 导出完成: {os.path.basename(json_path)}")
    except Exception as e:
        errors.append(f"[Exporter] JSON 导出失败: {e}")

    # 导出 CSV
    try:
        csv_path = export_csv(state, output_dir)
        paths["csv_path"] = csv_path
        progress.append(f"[Exporter] CSV 导出完成: {os.path.basename(csv_path)}")
    except Exception as e:
        errors.append(f"[Exporter] CSV 导出失败: {e}")

    # 导出 Excel
    try:
        excel_path = export_excel(state, output_dir)
        paths["excel_path"] = excel_path
        progress.append(f"[Exporter] Excel 导出完成: {os.path.basename(excel_path)}")
    except Exception as e:
        errors.append(f"[Exporter] Excel 导出失败: {e}")
        if "openpyxl" not in str(e):
            import traceback
            errors.append(traceback.format_exc())

    # 生成摘要
    summary = {
        "requirements_count": len(state.get("structured_requirements", [])),
        "blackbox_count": len(state.get("blackbox_tests", [])),
        "whitebox_count": len(state.get("whitebox_tests", [])),
        "optimized_count": len(state.get("optimized_suite", [])),
        "files": list(paths.values()),
    }
    paths["summary"] = summary

    progress.append(f"[Exporter] 所有产物导出完成！")

    return {
        "export_artifact": paths,
        "current_step": "completed",
        "completed_steps": state.get("completed_steps", []) + ["export_artifacts"],
        "progress_messages": progress,
        "errors": errors,
    }
